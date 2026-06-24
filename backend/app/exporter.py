from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.schemas import ReconcileRun


def _autosize(sheet) -> None:
    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 10), 28)


def _fills() -> dict[str, PatternFill]:
    return {
        "matched": PatternFill("solid", fgColor="DDEEDF"),
        "minor_diff": PatternFill("solid", fgColor="FFF2CC"),
        "major_diff": PatternFill("solid", fgColor="F4CCCC"),
        "failed": PatternFill("solid", fgColor="F4CCCC"),
    }


def build_run_excel(run: ReconcileRun) -> BytesIO:
    workbook = Workbook()

    summary_sheet = workbook.active
    summary_sheet.title = "横向汇总"
    sources = sorted({source for row in run.summary_rows for source in row.values})
    diff_sources = sorted({source for row in run.summary_rows for source in row.diffs})
    summary_sheet.append(["周期", "指标", *sources, *[f"差异-{source}" for source in diff_sources], "状态"])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)
    for row in run.summary_rows:
        summary_sheet.append(
            [
                row.period,
                row.metric,
                *[float(row.values.get(source, 0)) for source in sources],
                *[float(row.diffs.get(source, 0)) for source in diff_sources],
                row.status,
            ]
        )

    detail_sheet = workbook.create_sheet("店铺明细")
    detail_sheet.append(["周期", "店铺", "指标", "来源", "值", "基准值", "差异值", "差异率", "容差", "状态"])
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)

    fills = _fills()
    for row in run.rows:
        detail_sheet.append(
            [
                row.period,
                row.store,
                row.metric,
                row.source,
                float(row.value),
                float(row.warehouse_value),
                float(row.diff_value),
                float(row.diff_rate) if row.diff_rate is not None else None,
                float(row.tolerance),
                row.status,
            ]
        )
        fill = fills.get(row.status)
        if fill:
            for cell in detail_sheet[detail_sheet.max_row]:
                cell.fill = fill

    params_sheet = workbook.create_sheet("运行参数")
    params_sheet.append(["参数", "值"])
    params_sheet.append(["运行ID", run.id])
    params_sheet.append(["规则ID", run.rule_id])
    params_sheet.append(["规则名称", run.rule_name])
    params_sheet.append(["开始日期", run.start_date.isoformat()])
    params_sheet.append(["结束日期", run.end_date.isoformat()])
    params_sheet.append(["粒度", run.granularity])
    params_sheet.append(["状态", run.status])
    params_sheet.append(["错误信息", run.error_message or ""])
    for cell in params_sheet[1]:
        cell.font = Font(bold=True)

    for sheet in (summary_sheet, detail_sheet, params_sheet):
        _autosize(sheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def _run_label(run: ReconcileRun) -> str:
    return f"{run.rule_name} #{run.id}"


def build_compare_excel(runs: list[ReconcileRun]) -> BytesIO:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "所选结果对比"
    labels = [_run_label(run) for run in runs]
    summary_sheet.append(["周期", "指标", *labels])
    for cell in summary_sheet[1]:
        cell.font = Font(bold=True)

    grouped: dict[tuple[str, str], dict[str, float]] = {}
    for run in runs:
        label = _run_label(run)
        for row in run.summary_rows:
            key = (row.period, row.metric)
            grouped.setdefault(key, {})
            if row.values:
                grouped[key][label] = float(sum(row.values.values()))

    for (period, metric), values in sorted(grouped.items()):
        summary_sheet.append([period, metric, *[values.get(label, 0) for label in labels]])

    detail_sheet = workbook.create_sheet("店铺明细")
    detail_sheet.append(["周期", "店铺", "指标", "运行结果", "来源", "值", "差异值", "状态"])
    for cell in detail_sheet[1]:
        cell.font = Font(bold=True)
    fills = _fills()
    for run in runs:
        label = _run_label(run)
        for row in run.rows:
            detail_sheet.append(
                [row.period, row.store, row.metric, label, row.source, float(row.value), float(row.diff_value), row.status]
            )
            fill = fills.get(row.status)
            if fill:
                for cell in detail_sheet[detail_sheet.max_row]:
                    cell.fill = fill

    params_sheet = workbook.create_sheet("运行结果")
    params_sheet.append(["运行ID", "规则", "日期范围", "粒度", "状态", "创建时间"])
    for cell in params_sheet[1]:
        cell.font = Font(bold=True)
    for run in runs:
        params_sheet.append(
            [
                run.id,
                run.rule_name,
                f"{run.start_date.isoformat()} 至 {run.end_date.isoformat()}",
                run.granularity,
                run.status,
                run.created_at.isoformat(),
            ]
        )

    for sheet in (summary_sheet, detail_sheet, params_sheet):
        _autosize(sheet)

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream
